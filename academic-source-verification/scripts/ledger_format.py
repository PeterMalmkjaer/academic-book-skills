#!/usr/bin/env python3
"""
ledger_format.py — apply the house presentation convention to a reference-audit ledger.

Makes the emitted regneark readable (not just correct): a front 'Læsevejledning' legend
sheet, styled/frozen/filtered data sheets with tuned widths + wrapping + banding, and a
back 'Fejl fundet & rettet' report sheet. See references/regneark_layout.md.

NEVER writes to the frozen master — always reads --in and writes a separate --out copy.

Usage:
  python ledger_format.py --in ledger.xlsx --out ledger_pretty.xlsx [--errors errors.json]

errors.json (optional): {"report": "…", "rows": [[ "#","Type","Kilde/sted","Kapitel",
  "Hvad var forkert","Rettelse","Verificeret via","Konfidens","Delt m. DA?","DA-pendant",
  "RUN","Status" ], ...]}  # first row after header = row 1; header is added automatically.
"""
import argparse, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY, LIGHT, GREEN, RED = "1F4E79", "EAF1F8", "2E7D32", "C0392B"
HFILL = PatternFill("solid", fgColor=NAVY); HFONT = Font(color="FFFFFF", bold=True, size=11)
TITLE = Font(bold=True, size=15, color=NAVY); SUB = Font(bold=True, size=12, color=NAVY)
BAND = PatternFill("solid", fgColor=LIGHT); SUBFILL = PatternFill("solid", fgColor=LIGHT)
WRAP = Alignment(wrap_text=True, vertical="top"); TOP = Alignment(vertical="top")
CEN = Alignment(horizontal="center", vertical="center"); BOLD = Font(bold=True)
_thin = Side(style="thin", color="BFBFBF"); BORDER = Border(_thin, _thin, _thin, _thin)

# per-column widths + which columns wrap (by header NAME, robust to column order)
WIDTH = {"Kapitel":12,"Bib-nøgle":27,"Reference (forkortet)":24,"Tier":6,"p_correct":9,
         "Verificeringskommentar":58,"verifikationsmetoder":27,"antal_kilder":10,
         "kilde_evidens":44,"annotation_ref":16,"retraktionsstatus":15,"claim_støttet":13,
         "in_text_match":14,"verificeret_af":16,"verifikationsdato":15,"værktøj_version":24}
WRAPNAMES = {"Verificeringskommentar","verifikationsmetoder","kilde_evidens","værktøj_version"}

LEGEND_COLS = [
 ("Kapitel","hvilket kapitel/appendiks kilden citeres i"),
 ("Bib-nøgle","nøglen i references.bib (fx MellstromJohannesson2008)"),
 ("Reference (forkortet)","kort form: forfatter + år"),
 ("Tier","verifikationsniveau — T1 = fuldt verificeret (indhold + metadata); T2 = kun metadata"),
 ("p_correct","sandsynlighed i % (0–100) for at referencen er korrekt"),
 ("Verificeringskommentar","fritekst: hvad blev verificeret + hvilke kilder"),
 ("verifikationsmetoder","hvilke metoder er brugt (se koder nedenfor)"),
 ("antal_kilder","antal uafhængige, verificerbare kilder konsulteret"),
 ("kilde_evidens","de konkrete beviser: DOI'er, URL'er, bog-sider"),
 ("annotation_ref","peger på highlight/annotation lavet i artiklens PDF (cbs-libsearch)"),
 ("retraktionsstatus","er kilden trukket tilbage/korrigeret (se værdier)"),
 ("claim_støttet","siger kilden faktisk det, teksten påstår (se værdier)"),
 ("in_text_match","matcher inline/tabel-gengivelsen references.bib (se værdier)"),
 ("verificeret_af","hvem signerede (initialer) + om AI assisterede, fx 'PM (AI: ja)'"),
 ("verifikationsdato","dato for (seneste) verifikation, ÅÅÅÅ-MM-DD"),
 ("værktøj_version","værktøj + version brugt — til AI-/kildedeklarationen")]
LEGEND_METHODS = [("bog_læst / primærbog","kilden/primærkilden læst direkte"),
 ("referencer","tjekket mod references.bib / litteraturlisten"),("crossref/doi","DOI-resolver / CrossRef"),
 ("scite","scite (citationskontekst + retraktion)"),("elicit","Elicit"),("web","websøgning (fx Exa)"),
 ("retraction_watch","Retraction Watch"),("pdf_annotation","highlight/annotation i artiklens PDF")]
ERR_HDR = ["#","Type","Kilde / sted","Kapitel","Hvad var forkert","Rettelse","Verificeret via",
           "Konfidens","Delt m. DA?","DA-pendant","RUN","Status"]

def header_row(ws):
    for r in (1, 2):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() == "Kapitel":
            return r
    return None

def style_data_sheet(ws, hr):
    hdr = [c.value for c in ws[hr]]
    ncol = max((i for i, v in enumerate(hdr, 1) if v is not None), default=1)
    if hr == 2:
        ws.cell(row=1, column=1).font = SUB
    for c in range(1, ncol + 1):
        name = ws.cell(row=hr, column=c).value
        cell = ws.cell(row=hr, column=c); cell.fill = HFILL; cell.font = HFONT
        cell.alignment = CEN; cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = WIDTH.get(name, 16)
    for r in range(hr + 1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value in (None, "") and ws.cell(row=r, column=2).value in (None, ""):
            continue
        band = ((r - hr) % 2 == 0)
        for c in range(1, ncol + 1):
            name = ws.cell(row=hr, column=c).value
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP if name in WRAPNAMES else TOP
            cell.border = BORDER
            if band:
                cell.fill = BAND
    ws.freeze_panes = f"D{hr+1}"
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(ncol)}{ws.max_row}"

def build_legend(wb):
    if "Læsevejledning" in wb.sheetnames:
        del wb["Læsevejledning"]
    ws = wb.create_sheet("Læsevejledning", 0); ws.sheet_properties.tabColor = GREEN
    ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 95
    r = [1]
    def line(a, b="", style=None):
        ca = ws.cell(row=r[0], column=1, value=a); cb = ws.cell(row=r[0], column=2, value=b)
        ca.alignment = WRAP; cb.alignment = WRAP
        if style == "title": ca.font = TITLE
        elif style == "sub": ca.font = SUB; ca.fill = SUBFILL; cb.fill = SUBFILL
        elif style == "key": ca.font = BOLD
        r[0] += 1
    line("Reference-audit — hovedbog", "Læsevejledning, indhold og forkortelser", "title"); r[0] += 1
    line("Om filen", "Denne fil dokumenterer, at bogens kilder er efterprøvet — og hvordan. En fryst master (rør ikke) er kilden til bøgerne; denne er arbejdskopien.", "key"); r[0] += 1
    line("INDHOLD (faner)", "", "sub")
    line("Læsevejledning", "Denne side: indhold + forklaring af kolonner og forkortelser.")
    for ws2 in wb.worksheets:
        if ws2.title in ("Læsevejledning",): continue
        line(ws2.title, "")
    r[0] += 1
    line("KOLONNER", "", "sub")
    for k, v in LEGEND_COLS: line(k, v, "key")
    r[0] += 1
    line("KODER · verifikationsmetoder", "", "sub")
    for k, v in LEGEND_METHODS: line(k, v, "key")
    r[0] += 1
    line("VÆRDIER", "", "sub")
    line("retraktionsstatus", "none · retracted · corrected · erratum · EoC · ikke tjekket", "key")
    line("claim_støttet", "ja · delvist · nej · na · ikke tjekket", "key")
    line("in_text_match", "match · drift-rettet · na · ikke tjekket", "key")
    r[0] += 1
    line("BEMÆRK", "", "sub")
    line("'ikke tjekket'", "= kolonnen er endnu ikke udfyldt for den kilde. Et ÆRLIGT signal — ikke en påstand om, at den er tjekket.", "key")
    line("Provenans", "Provenans-kolonnerne er selve revisionssporet: HVORDAN hver kilde er efterprøvet — til forlagets/læserens transparens + AI-deklarationen.", "key")

def build_errors(wb, spec):
    if "Fejl fundet & rettet" in wb.sheetnames:
        del wb["Fejl fundet & rettet"]
    ws = wb.create_sheet("Fejl fundet & rettet"); ws.sheet_properties.tabColor = RED
    for c, w in {1:5,2:20,3:30,4:14,5:40,6:40,7:44,8:10,9:16,10:12,11:10,12:10}.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    er = 1
    ws.cell(row=er, column=1, value="Fejl fundet & rettet").font = TITLE; er += 2
    report = (spec or {}).get("report", "RAPPORT: (udfyld — mønster bag fejlene; at masteren er urørt).")
    cell = ws.cell(row=er, column=1, value=report); cell.alignment = WRAP
    ws.merge_cells(start_row=er, start_column=1, end_row=er, end_column=12)
    ws.row_dimensions[er].height = 64; er += 2
    for c, h in enumerate(ERR_HDR, 1):
        cell = ws.cell(row=er, column=c, value=h); cell.fill = HFILL; cell.font = HFONT
        cell.alignment = CEN; cell.border = BORDER
    hr = er; ws.freeze_panes = f"A{er+1}"; er += 1
    for row in (spec or {}).get("rows", []):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=er, column=c, value=v); cell.alignment = WRAP; cell.border = BORDER
            if (er - hr) % 2 == 0: cell.fill = BAND
        ws.row_dimensions[er].height = 46; er += 1
    if er > hr + 1:
        ws.auto_filter.ref = f"A{hr}:L{er-1}"

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--errors", help="optional errors.json (report + rows)")
    a = ap.parse_args()
    wb = openpyxl.load_workbook(a.inp)
    for ws in list(wb.worksheets):
        hr = header_row(ws)
        if hr is not None:
            style_data_sheet(ws, hr)
    if a.errors:
        spec = json.load(open(a.errors, encoding="utf-8"))
        build_errors(wb, spec)
    build_legend(wb)  # last, so the sheet index lists every other tab
    wb.save(a.out)
    print("wrote", a.out, "| sheets:", wb.sheetnames)

if __name__ == "__main__":
    main()
